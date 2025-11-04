# Copyright (C) 2025 NextERP Romania SRL
# License OPL-1.0 or later
# (https://www.odoo.com/documentation/user/19.0/legal/licenses/licenses.html#).

import base64
import io
import logging
from collections import defaultdict

from odoo import api, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)


class BomExcelImportWizard(models.TransientModel):
    _name = "bom.excel.import.wizard"
    _description = "BOM Excel Import Wizard - Two Step Process"

    # Step management
    step = fields.Selection(
        [
            ("upload", "Upload File"),
            ("operations", "Import Operations"),
            ("bom", "Import BOMs"),
            ("complete", "Import Complete"),
        ],
        default="upload",
        readonly=True,
    )

    # File upload fields
    excel_file = fields.Binary(required=True, help="Excel file containing BOM data")
    filename = fields.Char(required=True)

    # Options for import
    sheet_name = fields.Char(default="Sheet1", help="Name of the sheet to import from")
    start_row = fields.Integer(
        default=2,
        help="Row number to start importing from. Ussualy first is the column header.",
    )

    # Results and status
    operations_import_log = fields.Text(readonly=True)
    bom_import_log = fields.Text(readonly=True)
    operations_completed = fields.Boolean(default=False)
    bom_completed = fields.Boolean(default=False)

    # Summary fields
    workcenters_created = fields.Integer(readonly=True)
    operations_created = fields.Integer(readonly=True)
    boms_created = fields.Integer(readonly=True)
    bom_lines_created = fields.Integer(readonly=True)
    total_errors = fields.Integer(readonly=True)

    @api.model
    def default_get(self, fields_list):
        defaults = super().default_get(fields_list)
        return defaults

    def action_start_import(self):
        """Start the import process by validating file and moving to operations step"""
        if not openpyxl:
            raise UserError(
                self.env._(
                    "The openpyxl library is required to import Excel files. "
                    "Please install it with: pip install openpyxl"
                )
            )

        if not self.excel_file:
            raise UserError(self.env._("Please select an Excel file to import."))

        # Validate file format
        try:
            file_data = base64.b64decode(self.excel_file)
            file_like = io.BytesIO(file_data)
            workbook = openpyxl.load_workbook(file_like, data_only=True)

            # Check if sheet exists
            if self.sheet_name not in workbook.sheetnames:
                available_sheets = ", ".join(workbook.sheetnames)
                raise UserError(
                    self.env._(
                        "Sheet %(sheet_name)s not found in the Excel file. "
                        "Available sheets: %(available_sheets)s",
                        sheet_name=self.sheet_name,
                        available_sheets=available_sheets,
                    )
                )

        except Exception as e:
            raise UserError(
                self.env._("Error reading Excel file: %(error)s", error=str(e))
            ) from e

        # Move to operations step
        self.step = "operations"

        return {
            "type": "ir.actions.act_window",
            "res_model": "bom.excel.import.wizard",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_import_operations(self):
        """Import operations and workcenters from Excel file - Step 1"""
        try:
            # Decode the binary file
            file_data = base64.b64decode(self.excel_file)
            file_like = io.BytesIO(file_data)

            # Load workbook
            workbook = openpyxl.load_workbook(file_like, data_only=True)
            sheet = workbook[self.sheet_name]

            # Process the operations data
            results = self._process_operations_data(sheet)

            # Update the results
            self.operations_import_log = results["log"]
            self.workcenters_created = results["workcenters_created"]
            self.operations_created = results["operations_created"]
            self.boms_created = results[
                "boms_created"
            ]  # BOMs are created in step 1 now
            self.total_errors += results["errors"]
            self.operations_completed = True

            # Move to BOM step
            self.step = "bom"

            return {
                "type": "ir.actions.act_window",
                "res_model": "bom.excel.import.wizard",
                "res_id": self.id,
                "view_mode": "form",
                "target": "new",
            }

        except Exception as e:
            raise UserError(
                self.env._("Error importing operations: %(error)s", error=str(e))
            ) from e

    def action_import_boms(self):
        """Import BOMs from Excel file - Step 2"""
        try:
            # Decode the binary file
            file_data = base64.b64decode(self.excel_file)
            file_like = io.BytesIO(file_data)

            # Load workbook
            workbook = openpyxl.load_workbook(file_like, data_only=True)
            sheet = workbook[self.sheet_name]

            # Process the BOM data
            results = self._process_bom_data(sheet)

            # Update the results
            self.bom_import_log = results["log"]
            self.boms_created = results["boms_created"]
            self.bom_lines_created = results["bom_lines_created"]
            self.total_errors += results["errors"]
            self.bom_completed = True

            # Move to complete step
            self.step = "complete"

            return {
                "type": "ir.actions.act_window",
                "res_model": "bom.excel.import.wizard",
                "res_id": self.id,
                "view_mode": "form",
                "target": "new",
            }

        except Exception as e:
            raise UserError(
                self.env._("Error importing BOMs: %(error)s", error=str(e))
            ) from e

    def action_restart(self):
        """Restart the import process"""
        self.write(
            {
                "step": "upload",
                "operations_import_log": False,
                "bom_import_log": False,
                "operations_completed": False,
                "bom_completed": False,
                "workcenters_created": 0,
                "operations_created": 0,
                "boms_created": 0,
                "bom_lines_created": 0,
                "total_errors": 0,
            }
        )

        return {
            "type": "ir.actions.act_window",
            "res_model": "bom.excel.import.wizard",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def _process_operations_data(self, sheet):
        """Process the Excel sheet to create BOMs, workcenters and operations"""
        log_lines = []
        created_operations = 0
        created_workcenters = 0
        created_boms = 0
        errors = 0

        # Get existing data to avoid duplicates
        existing_workcenters = self.env["mrp.workcenter"].search([])  # pylint: disable=no-search-all
        existing_workcenter_names = {wc.name.lower() for wc in existing_workcenters}

        # Group data by product to create BOMs
        bom_data = defaultdict(list)
        workcenters_to_create = set()

        log_lines.append("=== OPERATIONS IMPORT LOG ===")
        log_lines.append(f"Processing sheet: {sheet.title}")
        log_lines.append("")

        # First pass: collect data and identify workcenters to create
        row_num = self.start_row
        for row in sheet.iter_rows(min_row=self.start_row, values_only=True):
            try:
                if not any(row):  # Skip empty rows
                    continue

                # Expected columns:
                # Column 1: Product produced
                # Column 2: Operation name
                # Column 3: Quantity consumed
                # Column 4: Product/component consumed UOM
                # Column 5: Product/component consumed
                # Column 6: Workcenter

                product_produced = row[0] if len(row) > 0 and row[0] else None
                operation_name = row[1] if len(row) > 1 and row[1] else None
                quantity = row[2] if len(row) > 2 and row[2] else None
                component_uom = row[3] if len(row) > 3 and row[3] else None
                component = row[4] if len(row) > 4 and row[4] else None
                workcenter_name = row[5] if len(row) > 5 and row[5] else None

                if not product_produced:
                    log_lines.append(f"Row {row_num}: Skipped - Missing product")
                    continue

                # Clean the data
                product_produced = str(product_produced).strip()
                operation_name = str(operation_name).strip() if operation_name else None
                component_uom = str(component_uom).strip() if component_uom else None
                workcenter_name = (
                    str(workcenter_name).strip() if workcenter_name else None
                )

                # Collect workcenters to create
                if (
                    workcenter_name
                    and workcenter_name.lower() not in existing_workcenter_names
                ):
                    workcenters_to_create.add(workcenter_name)

                # Group data by product
                bom_data[product_produced].append(
                    {
                        "operation": operation_name,
                        "workcenter": workcenter_name,
                        "component": component,
                        "component_uom": component_uom,
                        "quantity": quantity,
                        "row": row_num,
                    }
                )

                log_lines.append(
                    f"Row {row_num}: Processed - Product: {product_produced}, "
                    f"Operation: {operation_name}"
                )

            except Exception as e:
                errors += 1
                log_lines.append(f"Row {row_num}: ERROR - {str(e)}")
                _logger.error(f"Error processing row {row_num}: {str(e)}")

            row_num += 1

        log_lines.append("")
        log_lines.append("=== CREATING RECORDS ===")

        # Create workcenters first
        for workcenter_name in workcenters_to_create:
            try:
                self.env["mrp.workcenter"].create(
                    {
                        "name": workcenter_name,
                    }
                )
                created_workcenters += 1
                log_lines.append(f"Created workcenter: {workcenter_name}")
            except Exception as e:
                errors += 1
                log_lines.append(
                    f"ERROR creating workcenter {workcenter_name}: {str(e)}"
                )

        # Refresh workcenter cache
        all_workcenters = self.env["mrp.workcenter"].search([])  # pylint: disable=no-search-all
        workcenter_map = {wc.name: wc.id for wc in all_workcenters}

        # Get or create products and BOMs
        all_products = self.env["product.product"].search([])  # pylint: disable=no-search-all
        product_map = {p.name: p.id for p in all_products}
        product_map.update(
            {p.default_code: p.id for p in all_products if p.default_code}
        )

        # Create BOMs and operations for each product
        for product_name, operations_data in bom_data.items():
            try:
                # Find or create the main product
                product_id = self._find_or_create_product(
                    product_name, product_map, log_lines
                )
                if not product_id:
                    errors += 1
                    log_lines.append(
                        f"ERROR: Could not find or create product: {product_name}"
                    )
                    continue

                # Check if BOM already exists
                existing_bom = self.env["mrp.bom"].search(
                    [
                        (
                            "product_tmpl_id",
                            "=",
                            self.env["product.product"]
                            .browse(product_id)
                            .product_tmpl_id.id,
                        )
                    ],
                    limit=1,
                )

                if existing_bom:
                    log_lines.append(
                        f"WARNING: BOM already exists for {product_name}, skipping"
                    )
                    continue

                # Create BOM
                bom = self.env["mrp.bom"].create(
                    {
                        "product_tmpl_id": self.env["product.product"]
                        .browse(product_id)
                        .product_tmpl_id.id,
                        "product_id": product_id,
                        "type": "normal",
                        "product_qty": 1.0,
                    }
                )
                created_boms += 1
                log_lines.append(f"Created BOM for: {product_name}")

                # Create operations for this BOM
                sequence = 10
                unique_operations = {}
                for op_data in operations_data:
                    operation_name = op_data.get("operation")
                    workcenter_name = op_data.get("workcenter")

                    if operation_name and workcenter_name:
                        # Use operation name as key to avoid duplicates per BOM
                        if operation_name not in unique_operations:
                            if workcenter_name in workcenter_map:
                                try:
                                    operation = self.env[
                                        "mrp.routing.workcenter"
                                    ].create(
                                        {
                                            "name": operation_name,
                                            "bom_id": bom.id,
                                            "workcenter_id": workcenter_map[
                                                workcenter_name
                                            ],
                                            "sequence": sequence,
                                            "time_cycle_manual": 60,  # Default 1 hour
                                        }
                                    )
                                    unique_operations[operation_name] = operation.id
                                    created_operations += 1
                                    log_lines.append(
                                        f"  Created operation: {operation_name} -> "
                                        f"{workcenter_name}"
                                    )
                                    sequence += 10
                                except Exception as e:
                                    errors += 1
                                    log_lines.append(
                                        f"  ERROR creating operation {operation_name}: "
                                        f"{str(e)}"
                                    )
                            else:
                                errors += 1
                                log_lines.append(
                                    f"  ERROR: Workcenter {workcenter_name} not found "
                                    f"for operation {operation_name}"
                                )

            except Exception as e:
                errors += 1
                log_lines.append(f"ERROR creating BOM for {product_name}: {str(e)}")
                _logger.error(f"Error creating BOM for {product_name}: {str(e)}")

        log_lines.append("")
        log_lines.append("=== OPERATIONS SUMMARY ===")
        log_lines.append(f"BOMs created: {created_boms}")
        log_lines.append(f"Workcenters created: {created_workcenters}")
        log_lines.append(f"Operations created: {created_operations}")
        log_lines.append(f"Errors: {errors}")

        return {
            "log": "\n".join(log_lines),
            "workcenters_created": created_workcenters,
            "operations_created": created_operations,
            "boms_created": created_boms,
            "errors": errors,
        }

    def _process_bom_data(self, sheet):
        """Process the Excel sheet to add BOM lines to existing BOMs"""
        log_lines = []
        created_bom_lines = 0
        errors = 0

        # Group data by product
        bom_data = defaultdict(list)

        log_lines.append("=== BOM LINES IMPORT LOG ===")
        log_lines.append(f"Processing sheet: {sheet.title}")
        log_lines.append("")

        # First pass: collect and group BOM data
        row_num = self.start_row
        for row in sheet.iter_rows(min_row=self.start_row, values_only=True):
            try:
                if not any(row):  # Skip empty rows
                    continue

                # Expected columns:
                # Column 1: Product produced
                # Column 2: Operation name
                # Column 3: Quantity consumed
                # Column 4: Product/component consumed UOM
                # Column 5: Product/component consumed
                # Column 6: Workcenter

                product_produced = row[0] if len(row) > 0 and row[0] else None
                operation_name = row[1] if len(row) > 1 and row[1] else None
                quantity = row[2] if len(row) > 2 and row[2] else 1.0
                component_uom = row[3] if len(row) > 3 and row[3] else None
                component = row[4] if len(row) > 4 and row[4] else None
                _workcenter_name = row[5] if len(row) > 5 and row[5] else None

                if not product_produced or not component:
                    log_lines.append(
                        f"Row {row_num}: Skipped - Missing product or component"
                    )
                    continue

                # Clean and convert data
                product_produced = str(product_produced).strip()
                component = str(component).strip()
                try:
                    quantity = float(quantity) if quantity else 1.0
                except (ValueError, TypeError):
                    quantity = 1.0
                    log_lines.append(
                        f"Row {row_num}: Warning - Invalid quantity, using 1.0"
                    )

                operation_name = str(operation_name).strip() if operation_name else None
                component_uom = str(component_uom).strip() if component_uom else None

                # Add to BOM data
                bom_data[product_produced].append(
                    {
                        "component": component,
                        "component_uom": component_uom,
                        "quantity": quantity,
                        "operation": operation_name,
                        "row": row_num,
                    }
                )

                log_lines.append(
                    f"Row {row_num}: Collected - {product_produced} <- "
                    f" {quantity} x {component}"
                )

            except Exception as e:
                errors += 1
                log_lines.append(f"Row {row_num}: ERROR - {str(e)}")
                _logger.error(f"Error processing row {row_num}: {str(e)}")

            row_num += 1

        log_lines.append("")
        log_lines.append(f"Found {len(bom_data)} unique products to add BOM lines for")
        log_lines.append("")
        log_lines.append("=== ADDING BOM LINES ===")

        # Get product and operation mappings
        all_products = self.env["product.product"].search([])  # pylint: disable=no-search-all
        product_map = {p.name: p.id for p in all_products}
        product_map.update(
            {p.default_code: p.id for p in all_products if p.default_code}
        )

        # Get all existing BOMs
        all_boms = self.env["mrp.bom"].search([])  # pylint: disable=no-search-all
        bom_map = {}
        for bom in all_boms:
            if bom.product_id:
                bom_map[bom.product_id.name] = bom
                if bom.product_id.default_code:
                    bom_map[bom.product_id.default_code] = bom

        # Get all operations to link components
        all_operations = self.env["mrp.routing.workcenter"].search([])  # pylint: disable=no-search-all

        # Add BOM lines to existing BOMs
        for product_name, components in bom_data.items():
            try:
                # Find the BOM for this product
                if product_name not in bom_map:
                    log_lines.append(
                        f"WARNING: No BOM found for {product_name}, skipping"
                    )
                    continue

                bom = bom_map[product_name]
                log_lines.append(f"Adding components to BOM for: {product_name}")

                # Create BOM lines
                for comp_data in components:
                    comp_product_id = self._find_or_create_product(
                        comp_data["component"], product_map, log_lines
                    )
                    if not comp_product_id:
                        errors += 1
                        log_lines.append(
                            f"ERROR: Could not find component: {comp_data['component']}"
                        )
                        continue

                    # Find operation for this component if specified
                    operation_id = None
                    if comp_data.get("operation"):
                        operation = all_operations.filtered(
                            lambda op, bom=bom, comp=comp_data: op.name
                            == comp["operation"]
                            and op.bom_id.id == bom.id
                        )
                        if operation:
                            operation_id = operation[0].id

                    # Find UOM for the component
                    product_uom_id = None
                    if comp_data.get("component_uom"):
                        uom = self.env["uom.uom"].search(
                            [("name", "ilike", comp_data["component_uom"])], limit=1
                        )
                        if uom:
                            product_uom_id = uom.id
                        else:
                            log_lines.append(
                                f"  WARNING: UOM '{comp_data['component_uom']}' not "
                                f"found, using default"
                            )

                    # If no UOM found, use the product's default UOM
                    if not product_uom_id:
                        product = self.env["product.product"].browse(comp_product_id)
                        product_uom_id = product.uom_id.id

                    # Check if BOM line already exists for this component
                    existing_line = self.env["mrp.bom.line"].search(
                        [("bom_id", "=", bom.id), ("product_id", "=", comp_product_id)],
                        limit=1,
                    )

                    if existing_line:
                        log_lines.append(
                            f"  WARNING: Component {comp_data['component']} already "
                            f"exists in BOM, skipping"
                        )
                        continue

                    bom_line_vals = {
                        "bom_id": bom.id,
                        "product_id": comp_product_id,
                        "product_qty": comp_data["quantity"],
                        "product_uom_id": product_uom_id,
                    }

                    if operation_id:
                        bom_line_vals["operation_id"] = operation_id

                    self.env["mrp.bom.line"].create(bom_line_vals)
                    created_bom_lines += 1
                    operation_text = (
                        f" (Operation: {comp_data['operation']})"
                        if comp_data.get("operation")
                        else ""
                    )
                    uom_text = (
                        f" {comp_data['component_uom']}"
                        if comp_data.get("component_uom")
                        else ""
                    )
                    log_lines.append(
                        f"  Added component: {comp_data['quantity']}{uom_text} x "
                        f"{comp_data['component']}{operation_text}"
                    )

            except Exception as e:
                errors += 1
                log_lines.append(
                    f"ERROR adding components to BOM for {product_name}: {str(e)}"
                )
                _logger.error(
                    f"Error adding components to BOM for {product_name}: {str(e)}"
                )

        log_lines.append("")
        log_lines.append("=== BOM LINES SUMMARY ===")
        log_lines.append(f"BOM lines created: {created_bom_lines}")
        log_lines.append(f"Errors: {errors}")

        return {
            "log": "\n".join(log_lines),
            "boms_created": 0,  # BOMs were created in step 1
            "bom_lines_created": created_bom_lines,
            "errors": errors,
        }

    def _find_or_create_product(self, product_name, product_map, log_lines):
        """Find product by name or default_code, create if not found"""
        # Try to find existing product
        if product_name in product_map:
            return product_map[product_name]

        # Search by name (case insensitive)
        product = self.env["product.product"].search(
            [("name", "ilike", product_name)], limit=1
        )

        if product:
            product_map[product_name] = product.id
            return product.id

        # Search by default_code (case insensitive)
        product = self.env["product.product"].search(
            [("default_code", "ilike", product_name)], limit=1
        )

        if product:
            product_map[product_name] = product.id
            return product.id

        # Create new product
        try:
            categ = self.env.ref(
                "product.product_category_goods", raise_if_not_found=False
            )
            if not categ:
                categ = self.env["product.category"].search([], limit=1)
            product = self.env["product.product"].create(
                {
                    "name": product_name,
                    "default_code": product_name,
                    "type": "consu",  # Storable product
                    "is_storable": True,
                    "categ_id": categ.id,
                }
            )
            product_map[product_name] = product.id
            log_lines.append(f"Created product: {product_name} (storable, tracked)")
            return product.id
        except Exception as e:
            log_lines.append(f"ERROR creating product {product_name}: {str(e)}")
            return None
